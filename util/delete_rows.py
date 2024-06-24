import openpyxl

def delete_rows_after_condition(file_path):
    # Cargar el libro
    workbook = openpyxl.load_workbook(file_path)
    
    # Iterar a través de cada hoja
    for sheet in workbook.worksheets:
        print(sheet.title)
        # Inicializar una variable para guardar la fila desde donde eliminar
        delete_from_row = None

        # Iterar sobre las filas
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=3).value  # Columna C es la columna 3
            if cell_value is None or str(cell_value).strip().lower() == "fin" or "finn" in str(cell_value).strip().lower():
                delete_from_row = row
                break
        
        # Si encontramos una fila desde donde eliminar
        if delete_from_row:
            # Eliminar todas las filas desde delete_from_row hasta la última
            max_row = sheet.max_row
            sheet.delete_rows(delete_from_row, max_row - delete_from_row + 1)

    # Guardar el libro
    workbook.save(file_path)
    workbook.close()

# Ejemplo de uso
file_path = '..\\static\\cleansing_laos_ventas_2024.xlsx'
delete_rows_after_condition(file_path)
